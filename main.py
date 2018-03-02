import pymongo
from tornado.options import define, options
from tornado.web import RequestHandler
from tornado import gen
from openpyxl.chart.axis import DateAxis

import tornado.httpserver
import tornado.ioloop
import tornado.web

import motor
import threading
import subprocess
import random
import time
import os
import parseFile
import xlsxwriter
import openpyxl
from openpyxl.chart import *
import serial



define("port", default=8000, help="run on the given port", type=int)
cwd = os.getcwd() # used by static file server
print("Current Working directory :"+cwd)
parseFile.parse()
temperature = []
light = []
voltage = []
current = []
timestamp = []


class IndexPageHandler(RequestHandler):
    """
        index page
    """
    @gen.coroutine
    def get(self, *args, **kwargs):
        self.set_header('Content-Type', 'html')
        self.set_header("Access-Control-Allow-Origin", "*")
        self.set_header("Access-Control-Allow-Credentials", "false")
        self.set_header("Access-Control-Expose-Headers", "*")
        self.set_header("Access-Control-Allow-Methods", "*")
        self.set_header("Access-Control-Allow-Headers", "*")
        self.set_header("Access-Control-Allow-Headers", "accept, authorization")

        self.render('index.html')

    @gen.coroutine
    def post(self, *args, **kwargs):
        self.set_header('Content-Type', 'html')
        self.set_header("Access-Control-Allow-Origin", "*")
        self.set_header("Access-Control-Allow-Credentials", "false")
        self.set_header("Access-Control-Expose-Headers", "*")
        self.set_header("Access-Control-Allow-Methods", "*")
        self.set_header("Access-Control-Allow-Headers", "*")
        self.set_header("Access-Control-Allow-Headers", "accept, authorization")

        self.render('index.html')

# class LedPageHandler(RequestHandler):
#     """
#         Led page
#     """
#     @gen.coroutine
#     def get(self, *args, **kwargs):
#         self.set_header('Content-Type', 'html')
#         self.set_header("Access-Control-Allow-Origin", "*")
#         self.set_header("Access-Control-Allow-Credentials", "false")
#         self.set_header("Access-Control-Expose-Headers", "*")
#         self.set_header("Access-Control-Allow-Methods", "*")
#         self.set_header("Access-Control-Allow-Headers", "*")
#         self.set_header("Access-Control-Allow-Headers", "accept, authorization")
#
#         self.render('led.html')
#
#     @gen.coroutine
#     def post(self, *args, **kwargs):
#         self.set_header('Content-Type', 'html')
#         self.set_header("Access-Control-Allow-Origin", "*")
#         self.set_header("Access-Control-Allow-Credentials", "false")
#         self.set_header("Access-Control-Expose-Headers", "*")
#         self.set_header("Access-Control-Allow-Methods", "*")
#         self.set_header("Access-Control-Allow-Headers", "*")
#         self.set_header("Access-Control-Allow-Headers", "accept, authorization")
#
#         self.render('led.html')
#
# class DriverPageHandler(RequestHandler):
#     """
#         Driver page
#     """
#     @gen.coroutine
#     def get(self, *args, **kwargs):
#         self.set_header('Content-Type', 'html')
#         self.set_header("Access-Control-Allow-Origin", "*")
#         self.set_header("Access-Control-Allow-Credentials", "false")
#         self.set_header("Access-Control-Expose-Headers", "*")
#         self.set_header("Access-Control-Allow-Methods", "*")
#         self.set_header("Access-Control-Allow-Headers", "*")
#         self.set_header("Access-Control-Allow-Headers", "accept, authorization")
#
#         self.render('driver.html')
#
#     @gen.coroutine
#     def post(self, *args, **kwargs):
#         self.set_header('Content-Type', 'html')
#         self.set_header("Access-Control-Allow-Origin", "*")
#         self.set_header("Access-Control-Allow-Credentials", "false")
#         self.set_header("Access-Control-Expose-Headers", "*")
#         self.set_header("Access-Control-Allow-Methods", "*")
#         self.set_header("Access-Control-Allow-Headers", "*")
#         self.set_header("Access-Control-Allow-Headers", "accept, authorization")
#
#         self.render('driver.html')
#
# class DispatchedPageHandler(RequestHandler):
#     """
#         Dispatched page
#     """
#     @gen.coroutine
#     def get(self, *args, **kwargs):
#         self.set_header('Content-Type', 'html')
#         self.set_header("Access-Control-Allow-Origin", "*")
#         self.set_header("Access-Control-Allow-Credentials", "false")
#         self.set_header("Access-Control-Expose-Headers", "*")
#         self.set_header("Access-Control-Allow-Methods", "*")
#         self.set_header("Access-Control-Allow-Headers", "*")
#         self.set_header("Access-Control-Allow-Headers", "accept, authorization")
#
#         self.render('dispatched.html')
#
#     @gen.coroutine
#     def post(self, *args, **kwargs):
#         self.set_header('Content-Type', 'html')
#         self.set_header("Access-Control-Allow-Origin", "*")
#         self.set_header("Access-Control-Allow-Credentials", "false")
#         self.set_header("Access-Control-Expose-Headers", "*")
#         self.set_header("Access-Control-Allow-Methods", "*")
#         self.set_header("Access-Control-Allow-Headers", "*")
#         self.set_header("Access-Control-Allow-Headers", "accept, authorization")
#
#         self.render('dispatched.html')
#
# class OldGraphPageHandler(RequestHandler):
#     """
#         Old Graph page
#     """
#     @gen.coroutine
#     def get(self, *args, **kwargs):
#         self.set_header('Content-Type', 'html')
#         self.set_header("Access-Control-Allow-Origin", "*")
#         self.set_header("Access-Control-Allow-Credentials", "false")
#         self.set_header("Access-Control-Expose-Headers", "*")
#         self.set_header("Access-Control-Allow-Methods", "*")
#         self.set_header("Access-Control-Allow-Headers", "*")
#         self.set_header("Access-Control-Allow-Headers", "accept, authorization")
#
#         self.render('oldGraph.html')
#
#     @gen.coroutine
#     def post(self, *args, **kwargs):
#         self.set_header('Content-Type', 'html')
#         self.set_header("Access-Control-Allow-Origin", "*")
#         self.set_header("Access-Control-Allow-Credentials", "false")
#         self.set_header("Access-Control-Expose-Headers", "*")
#         self.set_header("Access-Control-Allow-Methods", "*")
#         self.set_header("Access-Control-Allow-Headers", "*")
#         self.set_header("Access-Control-Allow-Headers", "accept, authorization")
#
#         self.render('oldGraph.html')
#
# class CurrentGraphPageHandler(RequestHandler):
#     """
#         Current Graph page
#     """
#     @gen.coroutine
#     def get(self, *args, **kwargs):
#         self.set_header('Content-Type', 'html')
#         self.set_header("Access-Control-Allow-Origin", "*")
#         self.set_header("Access-Control-Allow-Credentials", "false")
#         self.set_header("Access-Control-Expose-Headers", "*")
#         self.set_header("Access-Control-Allow-Methods", "*")
#         self.set_header("Access-Control-Allow-Headers", "*")
#         self.set_header("Access-Control-Allow-Headers", "accept, authorization")
#
#         self.render('currentGraph.html')
#
#     @gen.coroutine
#     def post(self, *args, **kwargs):
#         self.set_header('Content-Type', 'html')
#         self.set_header("Access-Control-Allow-Origin", "*")
#         self.set_header("Access-Control-Allow-Credentials", "false")
#         self.set_header("Access-Control-Expose-Headers", "*")
#         self.set_header("Access-Control-Allow-Methods", "*")
#         self.set_header("Access-Control-Allow-Headers", "*")
#         self.set_header("Access-Control-Allow-Headers", "accept, authorization")
#
#         self.render('currentGraph.html')
#
# class ErrorHandler(RequestHandler):
#     """
#         404 Error page
#     """
#     @gen.coroutine
#     def get(self, *args, **kwargs):
#         self.set_header('Content-Type', 'html')
#         self.set_header("Access-Control-Allow-Origin", "*")
#         self.set_header("Access-Control-Allow-Credentials", "false")
#         self.set_header("Access-Control-Expose-Headers", "*")
#         self.set_header("Access-Control-Allow-Methods", "*")
#         self.set_header("Access-Control-Allow-Headers", "*")
#         self.set_header("Access-Control-Allow-Headers", "accept, authorization")
#
#         self.render("404.html")
#
#     @gen.coroutine
#     def post(self, *args, **kwargs):
#         self.set_header('Content-Type', 'html')
#         self.set_header("Access-Control-Allow-Origin", "*")
#         self.set_header("Access-Control-Allow-Credentials", "false")
#         self.set_header("Access-Control-Expose-Headers", "*")
#         self.set_header("Access-Control-Allow-Methods", "*")
#         self.set_header("Access-Control-Allow-Headers", "*")
#         self.set_header("Access-Control-Allow-Headers", "accept, authorization")
#
#         self.render("404.html")

class DataHandler(RequestHandler):
    """
        GIves New tempreature reading add to db
        :param localhost:8000/data
        :return list of new added temperature
    """
    @gen.coroutine
    def get(self, *args, **kwargs):
        self.set_header('Content-Type', 'application/json')
        self.set_header("Access-Control-Allow-Origin", "*")
        self.set_header("Access-Control-Allow-Credentials", "false")
        self.set_header("Access-Control-Expose-Headers", "*")
        self.set_header("Access-Control-Allow-Methods", "*")
        self.set_header("Access-Control-Allow-Headers", "*")
        self.set_header("Access-Control-Allow-Headers", "accept, authorization")



        temp = {"temp": temperature}
        lg = {"light": light}
        v = {"voltage": voltage}
        c = {"current": current}
        t = {"time": timestamp}
        data = {"Answer": [temp, lg, v, c, t],"status":200}
        #print(data)
        self.write(data)

class ResetGraphHandler(RequestHandler):
    """
        GIves New tempreature reading add to db
        :param localhost:8000/resetGraph
        :return list of new added temperature
    """
    @gen.coroutine
    def get(self, *args, **kwargs):
        self.set_header('Content-Type', 'application/json')
        self.set_header("Access-Control-Allow-Origin", "*")
        self.set_header("Access-Control-Allow-Credentials", "false")
        self.set_header("Access-Control-Expose-Headers", "*")
        self.set_header("Access-Control-Allow-Methods", "*")
        self.set_header("Access-Control-Allow-Headers", "*")
        self.set_header("Access-Control-Allow-Headers", "accept, authorization")

        global temperature, light, voltage, current, timestamp
        temperature = []
        light = []
        voltage = []
        current = []
        timestamp = []

        data = {"Answer": "","status":200}
        print(data)
        self.write(data)

class ExcelFileGenerator(RequestHandler):
    """
        GIves data from database about driver, led and dispatched and generates excel file
        :param localhost:8000/table
        :return Driver, led, dispatched data
    """

    @gen.coroutine
    def get(self, *args, **kwargs):
        self.set_header('Content-Type', 'application/json')
        self.set_header("Access-Control-Allow-Origin", "*")
        self.set_header("Access-Control-Allow-Credentials", "false")
        self.set_header("Access-Control-Expose-Headers", "*")
        self.set_header("Access-Control-Allow-Methods", "*")
        self.set_header("Access-Control-Allow-Headers", "*")
        self.set_header("Access-Control-Allow-Headers", "accept, authorization")

        db = motor.MotorClient('localhost').mullight
        try:


            driverdata = []
            result = db.driver.find({}, {"_id": 0}).sort('date', pymongo.DESCENDING)

            while (yield result.fetch_next):
                driverdata.append(result.next_object())

            leddata = []
            result = db.led.find({}, {"_id": 0}).sort('date', pymongo.DESCENDING)
            while (yield result.fetch_next):
                leddata.append(result.next_object())

            dispatchdata = []
            result = db.dispatched.find({}, {"_id": 0}).sort('date', pymongo.DESCENDING)
            while (yield result.fetch_next):
                dispatchdata.append(result.next_object())

            # data = {'Answer': [driver, led, dispatch],"status":200}

            wb = openpyxl.Workbook()
            ledworksheet = wb.create_sheet('Led',0)
            driverworksheet = wb.create_sheet('Driver',1)
            dispatchworksheet = wb.create_sheet('Dispatched', 2)
            graphworksheet = wb.create_sheet('Graphs', 3)

            # Add the worksheet data that the charts will refer to.
            ledworksheet.append(['Sr No', 'Model', 'Date'])
            for led in leddata:
                ledworksheet.append([led["sr_no"],led["model"],led["date"]])

            driverworksheet.append(['Sr No', 'Model', 'Date'])
            for driver in driverdata:
                driverworksheet.append([driver["sr_no"], driver["model"], driver["date"]])


            dispatchworksheet.append(['Led Sr No', 'Driver Sr No', 'Company', 'Model', 'Date', 'Status', 'Graph Data'])

            i=1
            d=2
            for dispatch in dispatchdata:

                dispatchworksheet.append([dispatch["l_sr_no"], dispatch["d_sr_no"], dispatch["company"], dispatch["model"], dispatch["date"], dispatch["status"],"Open Graph"])
                dispatchworksheet['G'+str(d)].value = '=HYPERLINK("#Graphs!A' + str(i) + '","Open Graph")'
                # dispatchworksheet.cell(d,7,'=HYPERLINK("#Graphs!A'+str(i)+'","Open Graph")')

                # graphworksheet = wb.active
                # graphworksheet.append([dispatch["l_sr_no"], dispatch["d_sr_no"]],i)
                graphworksheet.cell(i,1,dispatch["l_sr_no"])
                graphworksheet.cell(i,2,dispatch["d_sr_no"])
                graphworksheet.append(["Time","Tempreature","Light","Current","Voltage"])

                for j in range(0,len(dispatch["readings"][0]["temp"])):
                    graphworksheet.append([dispatch["readings"][4]["time"][j],dispatch["readings"][0]["temp"][j],dispatch["readings"][1]["light"][j],dispatch["readings"][2]["voltage"][j],dispatch["readings"][3]["current"][j]])

                # c1 = LineChart()
                # c1.title = "Line Chart"
                # c1.style = 13
                # c1.y_axis.title = 'Value'
                # c1.x_axis.title = 'Time'
                # data = Reference(graphworksheet, min_col=2, max_col=5, min_row=(i+2),max_row=len(dispatch["readings"][0]["temp"]))
                # c1.add_data(data, titles_from_data=True)
                # xdata = Reference(graphworksheet, min_col=1,min_row=(i + 3),max_row=len(dispatch["readings"][0]["temp"]))
                # c1.set_categories(xdata)
                #
                # graphworksheet.add_chart(c1,"G"+str(i+3))

                # if len(dispatch["readings"][0]["temp"]) > 15:
                i += len(dispatch["readings"][0]["temp"]) + 5
                # else:
                #     i += 20
                d +=1
                # print(dispatch['readings'])

            # dispatchworksheet['A1'] = "hello"
            # graphworksheet['B2'] = "goodbye"
            # dispatchworksheet['A1'].value = '=HYPERLINK("#graph!B2","Open Graph")'

            wb.save("mullight.xlsx")


            """
            workbook = xlsxwriter.Workbook('Mullight.xlsx')
            worksheetdriver = workbook.add_worksheet("Driver")
            worksheetled = workbook.add_worksheet("Led")
            worksheetdiepatched = workbook.add_worksheet("Dispatched")
            bold = workbook.add_format({'bold': 1})

            # Add the worksheet data that the charts will refer to.
            headings = ['Sr No', 'Model', 'Date']

            sr_no = []
            model= []
            date = []
            for driver in driverdata:
                sr_no.append(driver["sr_no"])
                model.append(driver["model"])
                date.append(driver["date"])

            worksheetdriver.write_row('A1', headings, bold)
            worksheetdriver.write_column('A2', sr_no)
            worksheetdriver.write_column('B2', model)
            worksheetdriver.write_column('C2', date)

            sr_no = []
            model = []
            date = []
            for driver in leddata:
                sr_no.append(driver["sr_no"])
                model.append(driver["model"])
                date.append(driver["date"])

            worksheetled.write_row('A1', headings, bold)
            worksheetled.write_column('A2', sr_no)
            worksheetled.write_column('B2', model)
            worksheetled.write_column('C2', date)

            headings = ['Led Sr No', 'Led Sr No', 'Company', 'Model', 'Date', 'Status']

            l_sr_no = []
            d_sr_no = []
            company = []
            status = []
            model = []
            date = []
            for driver in dispatchdata:
                l_sr_no.append(driver["l_sr_no"])
                d_sr_no.append(driver["d_sr_no"])
                company.append(driver["company"])
                status.append(driver["status"])
                model.append(driver["model"])
                date.append(driver["date"])

            worksheetdiepatched.write_row('A1', headings, bold)
            worksheetdiepatched.write_column('A2', l_sr_no)
            worksheetdiepatched.write_column('B2', d_sr_no)
            worksheetdiepatched.write_column('C2', company)
            worksheetdiepatched.write_column('D2', model)
            worksheetdiepatched.write_column('E2', date)
            worksheetdiepatched.write_column('F2', status)

            # Create a new chart object. In this case an embedded chart.
            # chart1 = workbook.add_chart({'type': 'line'})

            # Configure the first series.
            # chart1.add_series({
            #     'name': '=Sheet1!$B$1',
            #     'categories': '=Sheet1!$A$2:$A$7',
            #     'values': '=Sheet1!$B$2:$B$7',
            # })

            # Configure second series. Note use of alternative syntax to define ranges.
            # chart1.add_series({
            #     'name': ['Sheet1', 0, 2],
            #     'categories': ['Sheet1', 1, 0, 6, 0],
            #     'values': ['Sheet1', 1, 2, 6, 2],
            # })

            # Add a chart title and some axis labels.
            # chart1.set_title({'name': 'Results of sample analysis'})
            # chart1.set_x_axis({'name': 'Test number'})
            # chart1.set_y_axis({'name': 'Sample length (mm)'})

            # Set an Excel chart style. Colors with white outline and shadow.
            # chart1.set_style(10)

            # Insert the chart into the worksheet (with an offset).
            # worksheet.insert_chart('D2', chart1, {'x_offset': 25, 'y_offset': 10})

            workbook.close()

            # print(driver)

            # self.write(driver)
            """

            self.redirect("/Mullight.xlsx")

        except db:
            print("Answer", db)
            self.write('{"Answer":"fail","status":504}')

class ValidationHandler(RequestHandler):
    """
        Checks weather driver id and led id exits or not
        :param localhost:8000/check
        :return none or ok
    """
    @gen.coroutine
    def get(self, *args, **kwargs):
        self.set_header('Content-Type', 'application/json')
        self.set_header("Access-Control-Allow-Origin", "*")
        self.set_header("Access-Control-Allow-Credentials", "false")
        self.set_header("Access-Control-Expose-Headers", "*")
        self.set_header("Access-Control-Allow-Methods", "*")
        self.set_header("Access-Control-Allow-Headers", "*")
        self.set_header("Access-Control-Allow-Headers", "accept, authorization")

        try:
            db = motor.MotorClient('localhost').mullight

            driver_id = self.get_argument("driver")
            led_id = self.get_argument("led")

            d_result = yield db.driver.find_one({"sr_no":driver_id}, {"_id": 0})
            l_result = yield db.led.find_one({"sr_no": led_id}, {"_id": 0})

            if (d_result == None) or (l_result == None):
                self.write('{"Answer":"none","status":501}')
            else:
                self.write('{"Answer":"ok","status":500}')
        except:
            print("Answer", db.connect)
            self.write('{"Answer":"fail","status":504}')

class AddDevice(RequestHandler):
    """
        Add driver or led
        :param localhost:8000/addDevice
        :return none or ok
     """

    @gen.coroutine
    def get(self, *args, **kwargs):
        self.set_header('Content-Type', 'application/json')
        self.set_header("Access-Control-Allow-Origin", "*")
        self.set_header("Access-Control-Allow-Credentials", "false")
        self.set_header("Access-Control-Expose-Headers", "*")
        self.set_header("Access-Control-Allow-Methods", "*")
        self.set_header("Access-Control-Allow-Headers", "*")
        self.set_header("Access-Control-Allow-Headers", "accept, authorization")

        try:
            db = motor.MotorClient('localhost').mullight

            id = self.get_argument("id")
            model = self.get_argument("model")
            date = self.get_argument("date")
            dl = self.get_argument("dl")

            if date == "00-00-00":
                timestr = time.strftime("%d-%m-%y", time.localtime())
            else:
                timestr = date

            if (id == "" or id == "undefined" or model == "" or model == "undefined"):
                raise TypeError

            d_result = yield db.driver.find_one({"sr_no": id})
            l_result = yield db.led.find_one({"sr_no": id})
            dispatch_result = yield db.dispatched.find_one({"$or":[{"d_sr_no":id},{"l_sr_no":id}]})

            # print(dispatch_result)
            if dl == "true":
                if dispatch_result == None:
                    if d_result == None:
                        d_driver = yield db.driver.insert_one({"sr_no": id,"model":model,"date":timestr})
                        self.write('{"Answer":"ok","status":500}')

                    else:
                        self.write('{"Answer":"fail","status":501}')
                else:
                    self.write('{"Answer":"fail","status":501}')

            else :
                if dispatch_result == None:
                    if l_result == None:
                        d_led = yield db.led.insert_one({"sr_no": id,"model":model,"date":timestr})
                        self.write('{"Answer":"ok","status":500}')
                    else:
                        self.write('{"Answer":"fail","status":501}')
                else:
                    self.write('{"Answer":"fail","status":501}')

        except TypeError:
            print("Bad Request")
            self.write('{"Answer":"fail","status":202}')

        except:
            print("Add Device Handler:", end="")
            print("Answer", db.connect)
            self.write('{"Answer":"fail","status":504}')

class TestHandler(RequestHandler):
    """
        Checks weather driver id and led id exits or not
        :param localhost:8000/test
        :return none or ok
    """

    @gen.coroutine
    def get(self, *args, **kwargs):
        self.set_header('Content-Type', 'application/json')
        self.set_header("Access-Control-Allow-Origin", "*")
        self.set_header("Access-Control-Allow-Credentials", "false")
        self.set_header("Access-Control-Expose-Headers", "*")
        self.set_header("Access-Control-Allow-Methods", "*")
        self.set_header("Access-Control-Allow-Headers", "*")
        self.set_header("Access-Control-Allow-Headers", "accept, authorization")

        try:
            db = motor.MotorClient('localhost').mullight

            driver_id = self.get_argument("driver")
            led_id = self.get_argument("led")
            status = self.get_argument("status")
            company = self.get_argument("company")
            model = self.get_argument("model")
            entryexists = self.get_argument("valueexists")

            timestr = time.strftime("%d-%m-%y", time.localtime())

            if (company == "" or company == "undefined" or model == "" or model  == "undefined" or driver_id == "" or driver_id == "undefined" or led_id == "" or led_id == "undefined" or status == "" or status == "undefined" or entryexists == "" or entryexists == "undefined"):
                raise TypeError

            global temperature
            global light
            global current
            global voltage
            global timestamp

            l = []
            l.append({'temp':temperature})
            l.append({'light': light})
            l.append({'voltage': voltage})
            l.append({'current': current})
            l.append({'time': timestamp})

            # print(entryexists)
            if (entryexists == "false"):
                if (company == "" or company == "undefined"):
                    raise TypeError
                result = yield db.dispatched.find_one({"d_sr_no":driver_id})
                d_result = yield db.driver.find_one({"sr_no": driver_id})
                l_result = yield db.led.find_one({"sr_no": led_id})

                if result == None and d_result != None and l_result != None:
                    result = yield db.dispatched.insert_one({"d_sr_no": driver_id, "l_sr_no": led_id, "date": timestr, "readings": l, "status": status, "company": company, "model": model})
                    d_led = yield db.led.remove({"sr_no": led_id})
                    d_driver = yield db.driver.remove({"sr_no": driver_id})
                    if (d_led['ok'] == 1) and (d_driver['ok'] == 1):
                        temperature = []
                        light = []
                        voltage = []
                        current = []
                        timestamp = []
                        self.write('{"Answer":"ok","status":500}')
                    else:
                        self.write('{"Answer":"fail","status":503}')
                else:
                    self.write('{"Answer":"fail","status":501}')

            elif (entryexists == "true"):
                result = yield db.dispatched.find_one({"d_sr_no": driver_id})

                if result != None :
                    result = yield db.dispatched.update_one(
                        {"d_sr_no": driver_id, "l_sr_no": led_id}, { "$set": {"date": timestr, "readings": l, "status": status}})

                    # print(result)
                    temperature = []
                    light = []
                    voltage = []
                    current = []
                    timestamp = []
                    self.write('{"Answer":"ok","status":500}')
                else:
                    self.write('{"Answer":"fail","status":501}')

        except TypeError:
            print("Bad Request")
            self.write('{"Answer":"fail","status":202}')

        except:
            print("TestHandler :",end="")
            print("Answer", db.connect)
            self.write('{"Answer":"fail","status":504}')

class RDTestHandler(RequestHandler):
    """
        Checks weather driver id and led id exits or not
        :param localhost:8000/rdtest
        :return none or ok
    """

    @gen.coroutine
    def get(self, *args, **kwargs):
        self.set_header('Content-Type', 'application/json')
        self.set_header("Access-Control-Allow-Origin", "*")
        self.set_header("Access-Control-Allow-Credentials", "false")
        self.set_header("Access-Control-Expose-Headers", "*")
        self.set_header("Access-Control-Allow-Methods", "*")
        self.set_header("Access-Control-Allow-Headers", "*")
        self.set_header("Access-Control-Allow-Headers", "accept, authorization")

        try:

            db = motor.MotorClient('localhost').mullight

            driver_id = self.get_argument("driver")
            led_id = self.get_argument("led")
            status = self.get_argument("status")
            description = self.get_argument("description")
            model = self.get_argument("model")
            entryexists = self.get_argument("valueexists")

            print(driver_id)

            timestr = time.strftime("%d-%m-%y", time.localtime())

            if (description == "" or description == "undefined" or model == "" or model  == "undefined" or driver_id == "" or driver_id == "undefined" or led_id == "" or led_id == "undefined" or status == "" or status == "undefined" or entryexists == "" or entryexists == "undefined"):
                raise TypeError

            print(led_id)
            global temperature
            global light
            global current
            global voltage
            global timestamp

            l = []
            l.append({'temp':temperature})
            l.append({'light': light})
            l.append({'voltage': voltage})
            l.append({'current': current})
            l.append({'time': timestamp})

            # print(entryexists)

            if (entryexists == "false"):
                if (description == "" or description == "undefined"):
                    raise TypeError
                result = yield db.rd.find_one({"d_sr_no":driver_id,"l_sr_no":led_id,"description": description, "model": model})
                d_result = yield db.driver.find_one({"sr_no": driver_id})
                l_result = yield db.led.find_one({"sr_no": led_id})

                if result == None and d_result != None and l_result != None:
                    result = yield db.rd.insert_one({"d_sr_no": driver_id, "l_sr_no": led_id, "date": timestr, "readings": l, "status": status, "description": description, "model": model})
                    temperature = []
                    light = []
                    voltage = []
                    current = []
                    timestamp = []
                    self.write('{"Answer":"ok","status":500}')
                else:
                    self.write('{"Answer":"fail","status":501}')

            elif (entryexists == "true"):
                result = yield db.rd.find_one({"d_sr_no": driver_id,"l_sr_no":led_id,"description": description, "model": model})

                if result != None :
                    result = yield db.rd.update_one(
                        {"d_sr_no": driver_id, "l_sr_no": led_id,"description": description, "model": model}, { "$set": {"date": timestr, "readings": l, "status": status}})

                    # print(result)
                    temperature = []
                    light = []
                    voltage = []
                    current = []
                    timestamp = []
                    self.write('{"Answer":"ok","status":500}')
                else:
                    self.write('{"Answer":"fail","status":501}')

        except TypeError:
            print("Bad Request")
            self.write('{"Answer":"fail","status":202}')

        except:
            print("TestHandler :",end="")
            print("Answer", db.connect)
            self.write('{"Answer":"fail","status":504}')

class TableHandler(RequestHandler):
    """
        GIves data from database about driver, led and dispatched
        :param localhost:8000/table
        :return Driver, led, dispatched data
    """

    @gen.coroutine
    def get(self, *args, **kwargs):
        self.set_header('Content-Type', 'application/json')
        self.set_header("Access-Control-Allow-Origin", "*")
        self.set_header("Access-Control-Allow-Credentials", "false")
        self.set_header("Access-Control-Expose-Headers", "*")
        self.set_header("Access-Control-Allow-Methods", "*")
        self.set_header("Access-Control-Allow-Headers", "*")
        self.set_header("Access-Control-Allow-Headers", "accept, authorization")

        try:

            db = motor.MotorClient('localhost').mullight

            driver = []
            result = db.driver.find({}, {"_id": 0}).sort('date', pymongo.DESCENDING)

            while (yield result.fetch_next):
                driver.append(result.next_object())

            led = []
            result = db.led.find({}, {"_id": 0}).sort('date', pymongo.DESCENDING)
            while (yield result.fetch_next):
                led.append(result.next_object())

            dispatch = []
            result = db.dispatched.find({}, {"_id": 0}).sort('date', pymongo.DESCENDING)
            while (yield result.fetch_next):
                dispatch.append(result.next_object())

            data = {'Answer': [driver, led, dispatch],"status":200}
            self.write(data)
        except:
            print("Answer", db.connect)
            self.write('{"Answer":"fail","status":504}')

class RDTableHandler(RequestHandler):
    """
        GIves data from database about driver, led and dispatched
        :param localhost:8000/rdtable
        :return Driver, led, dispatched data
    """

    @gen.coroutine
    def get(self, *args, **kwargs):
        self.set_header('Content-Type', 'application/json')
        self.set_header("Access-Control-Allow-Origin", "*")
        self.set_header("Access-Control-Allow-Credentials", "false")
        self.set_header("Access-Control-Expose-Headers", "*")
        self.set_header("Access-Control-Allow-Methods", "*")
        self.set_header("Access-Control-Allow-Headers", "*")
        self.set_header("Access-Control-Allow-Headers", "accept, authorization")

        try:

            db = motor.MotorClient('localhost').mullight

            driver = []
            result = db.driver.find({}, {"_id": 0}).sort('date', pymongo.DESCENDING)

            while (yield result.fetch_next):
                driver.append(result.next_object())

            led = []
            result = db.led.find({}, {"_id": 0}).sort('date', pymongo.DESCENDING)
            while (yield result.fetch_next):
                led.append(result.next_object())

            rd = []
            result = db.rd.find({}, {"_id": 0}).sort('date', pymongo.DESCENDING)
            while (yield result.fetch_next):
                rd.append(result.next_object())

            data = {'Answer': [driver, led, rd],"status":200}
            self.write(data)
        except:
            print("Answer", db.connect)
            self.write('{"Answer":"fail","status":504}')

class MulLight(threading.Thread):

    def __init__(self,name):
        threading.Thread.__init__(self)
        self.name = name
        threading.Thread.start(self)

    def run(self):
        if self.name == "Server":
            print("-----Server Started-------")
            tornado.options.parse_command_line()
            app = tornado.web.Application(handlers=[(r"/", IndexPageHandler),
                                                    (r"/(.*\.html)", tornado.web.StaticFileHandler, {"path": cwd}),
                                                    (r"/(.*\.png)", tornado.web.StaticFileHandler, {"path": cwd}),
                                                    (r"/(.*\.jpg)", tornado.web.StaticFileHandler, {"path": cwd}),
                                                    (r"/(.*\.js)", tornado.web.StaticFileHandler, {"path": cwd}),
                                                    (r"/(.*\.css)", tornado.web.StaticFileHandler, {"path": cwd}),
                                                    (r"/data", DataHandler),
                                                    (r"/resetGraph", ResetGraphHandler),
                                                    (r"/check", ValidationHandler),
                                                    (r"/download", ExcelFileGenerator),
                                                    (r"/(.*\.xlsx)", tornado.web.StaticFileHandler, {"path": cwd}),
                                                    (r"/table", TableHandler),
                                                    (r"/rdtable", RDTableHandler),
                                                    (r"/test", TestHandler),
                                                    (r"/rdtest", RDTestHandler),
                                                    (r"/addDevice", AddDevice)
                                                    ])

            http_server = tornado.httpserver.HTTPServer(app)
            http_server.listen(options.port)
            tornado.ioloop.IOLoop.instance().start()
        elif (self.name == 'DB'):
            # subprocess.Popen(["C:\\Program Files\\MongoDB\\Server\\3.4\\bin\\mongod.exe", "--dbpath", "C:\data\db"])
            subprocess.call(r"C:\Program Files\MongoDB\Server\3.6\bin\mongod.exe --dbpath C:\data\db")
            print("Database closed")
        else :
            print("----Serial Reading------")
            #fromArduino = serial.Serial('COM3',9600)
            while True:
                temperature.append(random.randrange(32, 150))
                light.append((random.randrange(165, 20000)))
                voltage.append(random.randrange(34, 150))
                current.append(random.randrange(250, 5000))
                timestamp.append(time.strftime("%I:%M:%S", time.localtime()))
                # print(time.strftime("%I-%M-%S", time.localtime()))
                #print(temperature)
                time.sleep(10)


if __name__ == "__main__":
 
    print('Main')

    db = MulLight('DB')
    server = MulLight('Server')
    serial = MulLight('Serial Reader')


